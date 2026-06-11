"""Generate pricing-model.xlsx for Inbox-Agent.

Every cost/margin cell is a live Excel formula pointing at the Assumptions
sheet, so editing one input (token counts, provider rate, fixed cost) recalcs
all tiers and scenarios. Run: .pricing-venv/bin/python scripts/build_pricing_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- styling helpers -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # editable inputs = yellow
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")    # formula outputs = green
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0.0000'
MONEY2 = '$#,##0.00'
PCT = '0.0%'

def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

wb = Workbook()

# ===========================================================================
# 1. ASSUMPTIONS  (named cells so other sheets read them by reference)
# ===========================================================================
a = wb.active
a.title = "Assumptions"
a.column_dimensions["A"].width = 38
a.column_dimensions["B"].width = 16
a.column_dimensions["C"].width = 52

a["A1"] = "INBOX-AGENT PRICING MODEL — ASSUMPTIONS"
a["A1"].font = Font(bold=True, size=14, color="1F3864")
a["A2"] = "Yellow = editable input. Change these; every other sheet recalcs."
a["A2"].font = Font(italic=True, color="808080")

rows = [
    ("SECTION", "AI calls per email", "", ""),
    ("calls_per_email", 2, "Classifier + extractor (set to 1 if classifier is a heuristic)", "0"),
    ("SECTION", "Token sizes (per email, summed across calls)", "", ""),
    ("tok_cached_in", 1500, "System prompt, served from cache after 1st call", "#,##0"),
    ("tok_fresh_in", 1200, "Email body (HTML-stripped) + classifier hint", "#,##0"),
    ("tok_output", 250, "Structured JSON result", "#,##0"),
    ("SECTION", "Claude (claude-sonnet-4-6) rates  $/token", "", ""),
    ("rate_cache_read", 0.00000030, "$0.30 / M tokens", MONEY+'0000'),
    ("rate_input", 0.00000300, "$3.00 / M tokens", MONEY+'0000'),
    ("rate_output", 0.00001500, "$15.00 / M tokens", MONEY+'0000'),
    ("SECTION", "Gemini (gemini-2.0-flash) rates  $/token", "", ""),
    ("g_rate_input", 0.00000010, "$0.10 / M tokens", MONEY+'0000'),
    ("g_rate_output", 0.00000040, "$0.40 / M tokens", MONEY+'0000'),
    ("SECTION", "GCP fixed cost per portal", "", ""),
    ("fixed_portal_mo", 0.75, "Firestore storage, Secret Manager, idle Cloud Run — per portal/month", MONEY2),
    ("gcp_per_email", 0.0000020, "Cloud Run + Cloud Tasks + Firestore ops per email (~0)", MONEY+'0000'),
]

r = 4
named = {}
for key, val, note, fmt in rows:
    if key == "SECTION":
        a.cell(row=r, column=1, value=val).font = BOLD
        for c in range(1, 4):
            a.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
        continue
    a.cell(row=r, column=1, value=key)
    cell = a.cell(row=r, column=2, value=val)
    cell.fill = INPUT_FILL
    cell.border = BORDER
    cell.number_format = fmt
    a.cell(row=r, column=3, value=note).font = Font(italic=True, color="808080")
    named[key] = f"Assumptions!$B${r}"
    r += 1

# Derived per-email cost cells, also referenceable
r += 1
a.cell(row=r, column=1, value="DERIVED — per-email cost").font = BOLD
for c in range(1, 4):
    a.cell(row=r, column=c).fill = SECTION_FILL
r += 1
claude_email_row = r
a.cell(row=r, column=1, value="Claude cost / email")
a.cell(row=r, column=2,
       value=f"=({named['tok_cached_in']}*{named['rate_cache_read']}"
             f"+{named['tok_fresh_in']}*{named['rate_input']}"
             f"+{named['tok_output']}*{named['rate_output']})")
a.cell(row=r, column=2).number_format = MONEY
a.cell(row=r, column=2).fill = CALC_FILL
a.cell(row=r, column=3, value="Cached read + fresh input + output").font = Font(italic=True, color="808080")
named["claude_email"] = f"Assumptions!$B${r}"
r += 1
gem_email_row = r
a.cell(row=r, column=1, value="Gemini cost / email")
a.cell(row=r, column=2,
       value=f"=(({named['tok_cached_in']}+{named['tok_fresh_in']})*{named['g_rate_input']}"
             f"+{named['tok_output']}*{named['g_rate_output']})")
a.cell(row=r, column=2).number_format = MONEY
a.cell(row=r, column=2).fill = CALC_FILL
a.cell(row=r, column=3, value="Gemini has no separate cache tier; full input rate").font = Font(italic=True, color="808080")
named["gem_email"] = f"Assumptions!$B${r}"
r += 1
a.cell(row=r, column=1, value="Total marginal $/email (Claude + GCP)")
a.cell(row=r, column=2, value=f"={named['claude_email']}+{named['gcp_per_email']}")
a.cell(row=r, column=2).number_format = MONEY
a.cell(row=r, column=2).fill = CALC_FILL
named["marginal_email"] = f"Assumptions!$B${r}"

# ===========================================================================
# 2. TIERS
# ===========================================================================
t = wb.create_sheet("Tiers")
labels = ["Metric", "Starter", "Growth", "Pro", "Enterprise"]
widths = [34, 14, 14, 14, 16]
for i, w in enumerate(widths):
    t.column_dimensions[get_column_letter(i + 1)].width = w

t["A1"] = "PRICING TIERS — base + included bucket + overage"
t["A1"].font = Font(bold=True, size=13, color="1F3864")

style_header(t, 3, 5)
for i, lab in enumerate(labels):
    t.cell(row=3, column=i + 1, value=lab)

# input rows (editable)
base = [20, 79, 249, 750]
included = [100, 1000, 5000, 25000]
overage = [0.15, 0.10, 0.06, 0.035]

def setrow(row, label, vals, fmt=None, fill=None):
    t.cell(row=row, column=1, value=label).font = BOLD
    for i, v in enumerate(vals):
        cell = t.cell(row=row, column=i + 2, value=v)
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill

setrow(4, "Base price / month", base, MONEY2, INPUT_FILL)
setrow(5, "Included emails", included, '#,##0', INPUT_FILL)
setrow(6, "Overage $ / email", overage, MONEY, INPUT_FILL)

# formula rows — assume customer uses full bucket
mref = named["marginal_email"]
fref = named["fixed_portal_mo"]
for i in range(4):
    col = get_column_letter(i + 2)
    # COGS at full bucket = fixed + marginal*included
    t.cell(row=7, column=i + 2,
           value=f"={fref}+{col}5*{mref}").number_format = MONEY2
    t.cell(row=7, column=i + 2).fill = CALC_FILL
    t.cell(row=7, column=i + 2).border = BORDER
    # gross margin $ at base price = base - COGS
    t.cell(row=8, column=i + 2,
           value=f"={col}4-{col}7").number_format = MONEY2
    t.cell(row=8, column=i + 2).fill = CALC_FILL
    t.cell(row=8, column=i + 2).border = BORDER
    # gross margin %
    t.cell(row=9, column=i + 2,
           value=f"=IF({col}4=0,0,{col}8/{col}4)").number_format = PCT
    t.cell(row=9, column=i + 2).fill = CALC_FILL
    t.cell(row=9, column=i + 2).border = BORDER
    # effective $/email at cap
    t.cell(row=10, column=i + 2,
           value=f"={col}4/{col}5").number_format = MONEY
    t.cell(row=10, column=i + 2).fill = CALC_FILL
    t.cell(row=10, column=i + 2).border = BORDER
    # overage markup vs marginal cost
    t.cell(row=11, column=i + 2,
           value=f"={col}6/{mref}").number_format = '0.0"x"'
    t.cell(row=11, column=i + 2).fill = CALC_FILL
    t.cell(row=11, column=i + 2).border = BORDER

t.cell(row=7, column=1, value="COGS at full bucket (Claude)").font = BOLD
t.cell(row=8, column=1, value="Gross margin $ (base, full bucket)").font = BOLD
t.cell(row=9, column=1, value="Gross margin % (base, full bucket)").font = BOLD
t.cell(row=10, column=1, value="Effective $/email at cap").font = BOLD
t.cell(row=11, column=1, value="Overage markup vs cost").font = BOLD

t["A13"] = "Green = formula (recalcs from Assumptions). Yellow = editable."
t["A13"].font = Font(italic=True, color="808080")
t["A14"] = "Upgrade trigger: a tier is 'overflowing' once base + overages > next tier's base."
t["A14"].font = Font(italic=True, color="808080")

# ===========================================================================
# 3. PROVIDER SENSITIVITY — Claude vs Gemini at monthly volumes
# ===========================================================================
s = wb.create_sheet("Provider Sensitivity")
for i, w in enumerate([20, 18, 18, 18, 18]):
    s.column_dimensions[get_column_letter(i + 1)].width = w
s["A1"] = "PROVIDER COST SENSITIVITY — monthly AI+GCP cost by volume"
s["A1"].font = Font(bold=True, size=13, color="1F3864")
heads = ["Emails / month", "Claude cost", "Gemini cost", "Claude $/email", "Savings w/ Gemini"]
style_header(s, 3, 5)
for i, h in enumerate(heads):
    s.cell(row=3, column=i + 1, value=h)

vols = [100, 1000, 5000, 10000, 25000, 100000, 1000000]
cl = named["claude_email"]
gm = named["gem_email"]
gcp = named["gcp_per_email"]
fx = named["fixed_portal_mo"]
for j, v in enumerate(vols):
    row = 4 + j
    s.cell(row=row, column=1, value=v).number_format = '#,##0'
    s.cell(row=row, column=2, value=f"={fx}+A{row}*({cl}+{gcp})").number_format = MONEY2
    s.cell(row=row, column=3, value=f"={fx}+A{row}*({gm}+{gcp})").number_format = MONEY2
    s.cell(row=row, column=4, value=f"=B{row}/A{row}").number_format = MONEY
    s.cell(row=row, column=5, value=f"=B{row}-C{row}").number_format = MONEY2
    for c in range(2, 6):
        s.cell(row=row, column=c).fill = CALC_FILL
        s.cell(row=row, column=c).border = BORDER
    s.cell(row=row, column=1).border = BORDER

# ===========================================================================
# 4. CUSTOMER SCENARIO — pick a tier + actual volume, see the bill & margin
# ===========================================================================
c = wb.create_sheet("Customer Scenario")
for i, w in enumerate([34, 18, 52]):
    c.column_dimensions[get_column_letter(i + 1)].width = w
c["A1"] = "CUSTOMER SCENARIO CALCULATOR"
c["A1"].font = Font(bold=True, size=13, color="1F3864")
c["A2"] = "Edit the yellow inputs to model one customer's monthly bill & your margin."
c["A2"].font = Font(italic=True, color="808080")

def inp(row, label, val, fmt, note=""):
    c.cell(row=row, column=1, value=label).font = BOLD
    cell = c.cell(row=row, column=2, value=val)
    cell.fill = INPUT_FILL
    cell.number_format = fmt
    cell.border = BORDER
    if note:
        c.cell(row=row, column=3, value=note).font = Font(italic=True, color="808080")

def out(row, label, formula, fmt, note=""):
    c.cell(row=row, column=1, value=label).font = BOLD
    cell = c.cell(row=row, column=2, value=formula)
    cell.fill = CALC_FILL
    cell.number_format = fmt
    cell.border = BORDER
    if note:
        c.cell(row=row, column=3, value=note).font = Font(italic=True, color="808080")

inp(4, "Base price / month", 79, MONEY2, "what you charge this customer")
inp(5, "Included emails (bucket)", 1000, '#,##0')
inp(6, "Overage $ / email", 0.10, MONEY)
inp(7, "Actual emails processed", 1500, '#,##0', "their real monthly volume")

out(9, "Overage emails", "=MAX(0,B7-B5)", '#,##0')
out(10, "Overage charge", "=B9*B6", MONEY2)
out(11, "Customer pays (revenue)", "=B4+B10", MONEY2, "base + overage")
out(12, "Your COGS", f"={fx}+B7*({cl}+{gcp})", MONEY2, "fixed + Claude+GCP per actual email")
out(13, "Gross profit", "=B11-B12", MONEY2)
out(14, "Gross margin %", "=IF(B11=0,0,B13/B11)", PCT)
out(15, "Profit if run on Gemini", f"=B11-({fx}+B7*({gm}+{gcp}))", MONEY2, "same price, cheaper provider")

c["A17"] = "Tip: copy the Customer Scenario block per real client to build a portfolio view."
c["A17"].font = Font(italic=True, color="808080")

wb.save("pricing-model.xlsx")
print("wrote pricing-model.xlsx")
